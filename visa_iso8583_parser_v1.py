#!/usr/bin/env python3
"""
visa_iso8583_parser.py
=======================
VisaNet Authorization-Only ISO 8583 message parser.

Modeled on the NPSB w4_parser_v4.py architecture (same compact output
convention: 'FieldNum)value' joined by 'Ý', subfields joined by '³'),
but built against the VisaNet Authorization-Only Online Messages
Technical Specifications (Effective 20 April 2026).

Confirmed against live sample traffic (2026-07-05):
  - Standard 22-byte message header (H1-H12), reject header detection
    (H1 >= 26 and H13 bit1 == 1) is supported but untested against a
    live reject sample.
  - Primary bitmap bit1 = "secondary bitmap follows" FLAG, not field 1.
  - Secondary bitmap bit1 (=overall bit 65) = "tertiary bitmap follows"
    FLAG, not field 66. Field 66 starts at secondary-bitmap bit 2.
  - Switch deviation confirmed: some 0110 responses set the "secondary
    bitmap follows" flag to 1 even though the secondary bitmap content
    is all-zero and no F66-F128 field is present. Parser always
    consumes the 8-byte block when the flag is set, regardless of
    content, to avoid cascading byte-offset corruption.
  - Fixed-length BCD fields render ALL nibbles as-is (pad nibble is not
    stripped), e.g. F19)0050, F49)0050.
  - Variable-length BCD fields (F2, F35, F60, F62.2, etc.) DROP the
    leading pad nibble when length is odd; the length subfield gives
    the true digit count directly (per spec Table 11).
  - Masking convention: nibble 0xF -> '*' (redacted digit), nibble 0xD
    -> '=' (ISO 7813 Track 2 field separator). F52 (PIN block) is
    always fully masked (16 '*') since it is encrypted binary.
  - EBCDIC decode uses Python's built-in 'cp037' codec (confirmed
    against BANASREE MAIN R / DUTCH-BANGLA BANK PLC samples).
  - F34 / F104 / F111 use the Dataset-TLV structure (Length + DatasetID
    + DatasetLength + repeating Tag-Length-Value). F55 uses plain
    chained EMV TLV (no dataset wrapper).
  - F126 subfield semantics (.8/.9/.10) are only partially confirmed
    (raw hex passthrough for .8/.9); .10 EBCDIC+masking behavior is
    still unresolved pending more samples - flagged as TODO below.

KNOWN OPEN ITEMS (flag for more sample data before trusting fully):
  - Tertiary bitmap (fields 130-192) has not been observed in samples.
  - F17 (Date, Capture) attribute is "4 N" (EBCDIC digits), not BCD -
    unverified against a live sample.
  - F126 subfield-level structure beyond raw hex passthrough.
  - Reject message header (H13/H14) path is implemented per spec text
    only, not yet validated against live reject traffic.
"""

import re
import sys
from datetime import datetime

FIELD_SEP = "\u00dd"      # Ý  - separator between fields
SUBFIELD_SEP = "\u00b3"   # ³  - separator between repeated subfield groups

# ---------------------------------------------------------------------------
# Low level helpers
# ---------------------------------------------------------------------------

import re

def clean_hex(raw: str) -> bytes:
    hex_str = re.sub(r"\s+", "", raw)

    if len(hex_str) % 2:
        raise ValueError(
            f"Hex string has an odd number of characters ({len(hex_str)})"
        )

    return bytes.fromhex(hex_str)


def nibbles(data: bytes):
    """Yield each 4-bit nibble of a byte string, high nibble first."""
    for b in data:
        yield (b >> 4) & 0x0F
        yield b & 0x0F


NIBBLE_MASKED_MAP = {
    0xF: "*",   # redacted digit
    0xD: "D",   # track2 field separator - rendered literally as 'D', not ISO 7813 '='
}


def decode_bcd_full(data: bytes, masked: bool = False) -> str:
    """Decode ALL nibbles of a fixed-length BCD field (no pad stripping)."""
    out = []
    for n in nibbles(data):
        if masked and n in NIBBLE_MASKED_MAP:
            out.append(NIBBLE_MASKED_MAP[n])
        elif n <= 9:
            out.append(str(n))
        else:
            out.append("*")  # unexpected nibble - render as masked/unknown
    return "".join(out)


def decode_bcd_variable(data: bytes, length: int, masked: bool = False) -> str:
    """
    Decode a variable-length BCD field where `length` is the number of
    real digit *positions* (per the length subfield), dropping the
    leading pad nibble when length is odd.
    """
    nib = list(nibbles(data))
    if length % 2 != 0:
        nib = nib[1:]  # drop leading pad nibble
    nib = nib[:length]
    out = []
    for n in nib:
        if masked and n in NIBBLE_MASKED_MAP:
            out.append(NIBBLE_MASKED_MAP[n])
        elif n <= 9:
            out.append(str(n))
        else:
            out.append("*")
    return "".join(out)


def decode_ebcdic(data: bytes) -> str:
    try:
        return data.decode("cp037")
    except Exception:
        return data.hex().upper()


def decode_ebcdic_display(data: bytes) -> str:
    """EBCDIC decode, but replace non-printable control chars with a
    single space (confirmed against Field 44 subfield rendering, where
    control-byte 0x0F displays as a blank space rather than a glyph)."""
    s = decode_ebcdic(data)
    return "".join(c if (c.isprintable()) else " " for c in s)


def is_fully_printable_ebcdic(data: bytes) -> bool:
    try:
        s = data.decode("cp037")
    except Exception:
        return False
    return all(c.isprintable() for c in s)


def decode_bcd_bytes(data: bytes) -> str:
    """
    Decode a private-use variable BCD field where the length subfield is
    a BYTE count (not a digit-position count) - confirmed for Field 60
    (bit 60 is in Visa's private-use bit range 48, 60-63, 120-127).
    All nibbles are rendered directly with no pad-stripping.
    """
    return decode_bcd_full(data)


def to_hex(data: bytes) -> str:
    return data.hex().upper()


def bin_to_int(data: bytes) -> int:
    return int.from_bytes(data, "big")


# ---------------------------------------------------------------------------
# BER-TLV decoding
# ---------------------------------------------------------------------------

def read_tag(data: bytes, pos: int):
    first = data[pos]
    tag_bytes = [first]
    pos += 1
    if (first & 0x1F) == 0x1F:  # low 5 bits all 1 -> multi-byte tag
        while True:
            b = data[pos]
            tag_bytes.append(b)
            pos += 1
            if not (b & 0x80):
                break
    return bytes(tag_bytes).hex(), pos


def read_length(data: bytes, pos: int):
    first = data[pos]
    pos += 1
    if first & 0x80:  # long form
        num_bytes = first & 0x7F
        length = int.from_bytes(data[pos:pos + num_bytes], "big")
        pos += num_bytes
    else:  # short form
        length = first & 0x7F
    return length, pos


def render_int_or_empty(tag: str, value: bytes) -> str:
    n = int.from_bytes(value, "big")
    return "" if n == 0 else str(n)


def render_text_or_hex(tag: str, value: bytes) -> str:
    """Prefer EBCDIC text if fully printable, else fall back to hex."""
    if is_fully_printable_ebcdic(value):
        return decode_ebcdic(value)
    return to_hex(value)


def render_text_or_hex_dataset(dataset_id: str, tag: str, value: bytes) -> str:
    return render_text_or_hex(tag, value)


def render_hex(tag: str, value: bytes) -> str:
    return to_hex(value)


# Tags in Field 55 (ICC data) confirmed to render as plain decimal integers
# rather than hex (e.g. 9F27=CID single byte 0x80 -> "128", not "80").
F55_DECIMAL_TAGS = {"9f27", "9f33", "82"}


def render_f55_value(tag: str, value: bytes) -> str:
    if tag in F55_DECIMAL_TAGS:
        return render_int_or_empty(tag, value)
    return render_hex(tag, value)  # default: hex passthrough (confirmed for most tags)


# Field 34 dataset '01' (terminal/app version, IP address) confirmed as
# EBCDIC text; dataset '56' confirmed as decimal-int-with-zero-as-empty.
# Any other dataset ID defaults to hex passthrough pending more samples.
def render_f34_value(dataset_id: str, tag: str, value: bytes) -> str:
    if dataset_id == "01":
        return render_text_or_hex(tag, value)
    if dataset_id == "56":
        return render_int_or_empty(tag, value)
    return render_hex(tag, value)


def decode_plain_tlv(data: bytes, value_renderer=render_hex) -> dict:
    """Chained Tag-Length-Value elements with no dataset wrapper (e.g. F55)."""
    out = {}
    pos = 0
    n = len(data)
    while pos < n:
        tag, pos = read_tag(data, pos)
        length, pos = read_length(data, pos)
        value = data[pos:pos + length]
        pos += length
        out[tag] = value_renderer(tag, value)
    return out


def decode_dataset_tlv(data: bytes, value_renderer=None):
    """
    Dataset-TLV structure used by F34 / F104 / F111 / F123:
        DatasetID (1 byte) + DatasetLength (2 bytes) +
        repeating [Tag-Length-Value] within that dataset.
    The outer field's own Length prefix is assumed already stripped
    by the caller. Returns list of (dataset_id_hex, {tag: rendered}).
    `value_renderer(dataset_id, tag, value_bytes) -> str` if provided,
    else defaults to hex passthrough.
    """
    out = []
    pos = 0
    n = len(data)
    while pos < n:
        dataset_id = data[pos]
        pos += 1
        if pos + 2 > n:
            break
        dataset_len = int.from_bytes(data[pos:pos + 2], "big")
        pos += 2
        dataset_body = data[pos:pos + dataset_len]
        pos += dataset_len
        ds_id_str = f"{dataset_id:02x}"
        if value_renderer is not None:
            renderer = lambda tag, val, _ds=ds_id_str: value_renderer(_ds, tag, val)
        else:
            renderer = render_hex
        tlvs = decode_plain_tlv(dataset_body, value_renderer=renderer)
        out.append((ds_id_str, tlvs))
    return out


# ---------------------------------------------------------------------------
# Header parsing
# ---------------------------------------------------------------------------

def parse_header(data: bytes):
    h = {}
    pos = 0
    h["H1"] = data[pos]; pos += 1
    h["H2"] = data[pos]; pos += 1
    h["H3"] = data[pos]; pos += 1
    h["H4"] = bin_to_int(data[pos:pos + 2]); pos += 2
    h["H5"] = decode_bcd_full(data[pos:pos + 3]); pos += 3
    h["H6"] = decode_bcd_full(data[pos:pos + 3]); pos += 3
    h["H7"] = data[pos]; pos += 1
    h["H8"] = bin_to_int(data[pos:pos + 2]); pos += 2
    h["H9"] = bin_to_int(data[pos:pos + 3]); pos += 3
    h["H10"] = data[pos]; pos += 1
    h["H11"] = bin_to_int(data[pos:pos + 3]); pos += 3
    h["H12"] = data[pos]; pos += 1

    is_reject = False
    if h["H1"] >= 26 and pos < len(data):
        # peek ahead: H13 bitmap present only in reject headers
        h13_bit1 = data[pos] & 0x80
        if h13_bit1:
            is_reject = True
            h["H13"] = bin_to_int(data[pos:pos + 2]); pos += 2
            h["H14"] = decode_bcd_full(data[pos:pos + 2]); pos += 2

    h["_is_reject"] = is_reject
    h["_header_len"] = pos
    return h, pos


# ---------------------------------------------------------------------------
# Bitmap parsing
# ---------------------------------------------------------------------------

def parse_bitmap_block(data: bytes, pos: int, field_offset: int):
    """
    Parse one 8-byte bitmap block. Bit 1 is a FLAG (not a field), bits
    2-64 map to fields (field_offset+1) .. (field_offset+63).
    Returns (dict_of_present_fields->True, flag_bit1, next_pos).
    """
    block = data[pos:pos + 8]
    pos += 8
    bits = []
    for byte in block:
        for i in range(8):
            bits.append((byte >> (7 - i)) & 1)
    flag = bits[0]
    present = {}
    for i in range(1, 64):
        if bits[i]:
            present[field_offset + i] = True
    return present, flag, pos


def parse_all_bitmaps(data: bytes, pos: int):
    present = {}
    primary, flag2, pos = parse_bitmap_block(data, pos, field_offset=1)   # fields 2-64
    present.update(primary)
    if flag2:
        secondary, flag3, pos = parse_bitmap_block(data, pos, field_offset=65)  # fields 66-128
        present.update(secondary)
        if flag3:
            tertiary, _flag4, pos = parse_bitmap_block(data, pos, field_offset=129)  # fields 130-192
            present.update(tertiary)
    return present, pos


# ---------------------------------------------------------------------------
# Field definitions
# ---------------------------------------------------------------------------
# type: 'bcd_f' fixed BCD (full nibble render) | 'bcd_v' variable BCD (length-stripped)
#       'an' EBCDIC alnum fixed | 'ans_v' EBCDIC variable | 'bin' raw hex fixed
#       'special' -> custom handler by field number

FIELD_DEFS = {
    2:  {"type": "bcd_v",  "mask": True},                       # PAN
    3:  {"type": "bcd_f",  "len": 3},                           # Processing code
    4:  {"type": "bcd_f",  "len": 6},                           # Amount, transaction
    5:  {"type": "bcd_f",  "len": 6},                           # Amount, settlement
    6:  {"type": "bcd_f",  "len": 6},                           # Amount, cardholder billing
    7:  {"type": "bcd_f",  "len": 5},                           # Transmission date/time
    8:  {"type": "bcd_f",  "len": 4},                           # Amount, cardholder billing fee
    9:  {"type": "bcd_f",  "len": 4},                           # Conversion rate, settlement
    10: {"type": "bcd_f",  "len": 4},                           # Conversion rate, cardholder billing
    11: {"type": "bcd_f",  "len": 3},                           # STAN
    12: {"type": "bcd_f",  "len": 3},                           # Time, local
    13: {"type": "bcd_f",  "len": 2},                           # Date, local
    14: {"type": "bcd_f",  "len": 2},                           # Date, expiration
    15: {"type": "bcd_f",  "len": 2},                           # Date, settlement
    16: {"type": "bcd_f",  "len": 2},                           # Date, conversion
    17: {"type": "an",     "len": 4},                           # Date, capture (EBCDIC digits per spec)
    18: {"type": "bcd_f",  "len": 2},                           # Merchant type
    19: {"type": "bcd_f",  "len": 2},                           # Acquiring inst country code
    20: {"type": "bcd_f",  "len": 2},                           # PAN extended country code
    22: {"type": "bcd_f",  "len": 2},                           # POS entry mode
    23: {"type": "bcd_f",  "len": 2},                           # Card sequence number
    24: {"type": "bcd_f",  "len": 2},                           # NII
    25: {"type": "bcd_f",  "len": 1},                           # POS condition code
    26: {"type": "bcd_f",  "len": 1},                           # POS PIN capture code
    28: {"type": "special","len": 9},                           # Amount, transaction fee (1 AN sign + 8 N)
    32: {"type": "bcd_v"},                                      # Acquiring inst ID code
    33: {"type": "bcd_v"},                                      # Forwarding inst ID code
    34: {"type": "tlv_dataset", "len_bytes": 2},                # Acceptance environment data (confirmed 2-byte outer length)
    35: {"type": "track2"},                                     # Track 2 data
    36: {"type": "bcd_v"},                                      # Track 3 data
    37: {"type": "an",     "len": 12},                          # RRN
    38: {"type": "an",     "len": 6},                           # Auth ID response
    39: {"type": "an",     "len": 2},                           # Response code
    41: {"type": "ans",    "len": 8},                           # Card acceptor terminal ID
    42: {"type": "ans",    "len": 15},                          # Card acceptor ID code
    43: {"type": "ans",    "len": 40},                          # Card acceptor name/location
    44: {"type": "special_44"},                                 # Additional response data (subfields)
    45: {"type": "ans_v"},                                      # Track 1 data
    46: {"type": "ans_v"},                                      # Amounts, fees
    47: {"type": "ans_v"},                                      # Additional data, national
    48: {"type": "ans_v"},                                      # Additional data, private
    49: {"type": "bcd_f",  "len": 2},                           # Currency code, transaction
    50: {"type": "bcd_f",  "len": 2},                           # Currency code, settlement
    51: {"type": "bcd_f",  "len": 2},                           # Currency code, cardholder billing
    52: {"type": "pin"},                                        # PIN data (always masked)
    53: {"type": "bcd_f",  "len": 8},                           # Security related control info
    54: {"type": "ans_v"},                                      # Additional amounts
    55: {"type": "tlv_plain_v"},                                # ICC related data (EMV)
    56: {"type": "ans_v"},                                      # PAR data
    57: {"type": "ans_v"},                                      # Reserved national
    58: {"type": "ans_v"},                                      # Reserved national
    59: {"type": "ans_v"},                                      # National POS geographic data
    60: {"type": "bcd_v_bytes"},                                # Additional POS info (private-use bit)
    61: {"type": "bcd_v_bytes"},                                # Other amounts (private-use bit)
    62: {"type": "field62"},                                    # CPS fields (bitmap + subfields)
    63: {"type": "field63"},                                    # Private-use fields (bitmap + subfields)
    66: {"type": "bcd_f",  "len": 1},                           # Settlement code
    67: {"type": "bcd_f",  "len": 1},                           # Extended payment code
    68: {"type": "bcd_f",  "len": 2},                           # Receiving inst country code
    69: {"type": "bcd_f",  "len": 2},                           # Settlement inst country code
    70: {"type": "bcd_f",  "len": 2},                           # Network mgmt info code
    90: {"type": "bcd_f",  "len": 21},                          # Original data elements
    100:{"type": "bcd_v"},                                      # Receiving institution ID code
    102:{"type": "ans_v"},                                      # Account ID 1
    103:{"type": "ans_v"},                                      # Account ID 2
    104:{"type": "tlv_dataset", "len_bytes": 1},                # Transaction description (confirmed 1-byte outer length)
    111:{"type": "tlv_dataset", "len_bytes": 2},                # Amount, currency (confirmed 2-byte outer length)
    123:{"type": "tlv_dataset", "len_bytes": 1},                # Additional data 2 (unconfirmed - assumed same pattern as 104/111)
    124:{"type": "ans_v"},                                      # Additional data 3
    125:{"type": "ans_v"},                                      # Additional data 4
    126:{"type": "field126"},                                   # VSDC/issuer script - partially resolved
    127:{"type": "tlv_dataset", "len_bytes": 1},                # File maintenance/network data (unconfirmed)
    134:{"type": "bcd_v"},                                      # Visa discretionary data
    142:{"type": "bcd_v"},                                      # Issuer script
}

FIELD62_SUBFIELDS = {
    1: 1, 2: 8, 3: 4, 4: 1, 5: 1, 6: 1, 7: 26, 8: 3, 9: 1, 10: 3,
    11: 1, 12: 1, 13: 1, 14: 6, 15: 1, 16: 2, 17: 15, 18: 1, 19: 2,
    20: 5, 21: 4, 22: 6, 23: 2, 24: 6, 25: 1,
}
FIELD62_TYPE = {
    2: "bcd", 5: "bcd", 8: "bcd", 10: "bcd", 11: "bcd", 12: "bcd",
    16: "bcd", 20: "bcd",
}  # default 'an' otherwise

FIELD63_SUBFIELDS = {
    1: 2, 2: 2, 3: 2, 4: 2, 6: 7, 7: 8, 8: 4, 9: 14, 10: 13, 11: 1,
    12: 30, 13: 3, 14: 36, 15: 8, 18: 2, 19: 3, 21: 1,
}
FIELD63_TYPE = {1: "bcd", 2: "bcd", 3: "bcd", 4: "bcd", 8: "bcd", 18: "bcd"}


# ---------------------------------------------------------------------------
# Field decoding dispatcher
# ---------------------------------------------------------------------------

def read_len_prefix(data: bytes, pos: int, nbytes: int = 1):
    length = int.from_bytes(data[pos:pos + nbytes], "big")
    return length, pos + nbytes


def decode_field(fnum: int, data: bytes, pos: int):
    """Decode one data field starting at pos. Returns (rendered_str, new_pos)."""
    fdef = FIELD_DEFS.get(fnum)
    if fdef is None:
        # Unknown field: assume 1-byte length + raw hex passthrough (safe default)
        length, pos2 = read_len_prefix(data, pos)
        val = data[pos2:pos2 + length]
        return to_hex(val), pos2 + length

    t = fdef["type"]

    if t == "bcd_f":
        n = fdef["len"]
        val = decode_bcd_full(data[pos:pos + n])
        return val, pos + n

    if t == "an":
        n = fdef["len"]
        val = decode_ebcdic(data[pos:pos + n])
        return val, pos + n

    if t == "ans":
        n = fdef["len"]
        val = decode_ebcdic(data[pos:pos + n])
        return val, pos + n

    if t == "ans_v":
        length, pos2 = read_len_prefix(data, pos)
        val = decode_ebcdic(data[pos2:pos2 + length])
        return val, pos2 + length

    if t == "bcd_v":
        length, pos2 = read_len_prefix(data, pos)
        nbytes = (length + 1) // 2
        val = decode_bcd_variable(data[pos2:pos2 + nbytes], length, masked=fdef.get("mask", False))
        return val, pos2 + nbytes

    if t == "bcd_v_bytes":
        length, pos2 = read_len_prefix(data, pos)  # length = BYTE count (private-use field)
        val = decode_bcd_bytes(data[pos2:pos2 + length])
        return val, pos2 + length

    if t == "pin":
        # Field 52: fixed 8 bytes, always fully masked (encrypted PIN block)
        pos2 = pos + 8
        return "*" * 16, pos2

    if t == "track2":
        length, pos2 = read_len_prefix(data, pos)
        nbytes = (length + 1) // 2
        val = decode_bcd_variable(data[pos2:pos2 + nbytes], length, masked=True)
        return val, pos2 + nbytes

    if t == "special":  # Field 28 style: 1 AN sign + 8 N
        n = fdef["len"]
        raw = data[pos:pos + n]
        sign = decode_ebcdic(raw[0:1])
        amt = decode_bcd_full(raw[1:])
        return f"{sign}{amt}", pos + n

    if t == "special_44":
        return decode_field_44(data, pos)

    if t == "field62":
        return decode_field_62(data, pos)

    if t == "field63":
        return decode_field_63(data, pos)

    if t == "tlv_plain_v":
        # Confirmed structure for F55: 1-byte OUTER length (total bytes
        # that follow), then a 1-byte marker (constant 0x01 in both
        # confirmed samples), then a 2-byte INNER length for the actual
        # EMV TLV chain, then the chain itself.
        outer_len, pos2 = read_len_prefix(data, pos, nbytes=1)
        inner_len = int.from_bytes(data[pos2 + 1:pos2 + 3], "big")
        chain_start = pos2 + 3
        body = data[chain_start:chain_start + inner_len]
        renderer = render_f55_value if fnum == 55 else render_hex
        tlvs = decode_plain_tlv(body, value_renderer=renderer)
        parts = [f"F{fnum}.{tag})" + val for tag, val in tlvs.items()]
        return (SUBFIELD_SEP.join(parts) + SUBFIELD_SEP if parts else ""), pos2 + outer_len

    if t == "tlv_dataset":
        len_bytes = fdef.get("len_bytes", 1)
        length, pos2 = read_len_prefix(data, pos, nbytes=len_bytes)
        body = data[pos2:pos2 + length]
        if fnum == 34:
            renderer = render_f34_value
        elif fnum == 111:
            # Confirmed: F111 datasets carry genuine EBCDIC text
            # (e.g. "Visa Inc., Exchange Rate").
            renderer = render_text_or_hex_dataset
        else:
            # F104 confirmed hex passthrough (e.g. tag 83 -> "6196" is the
            # raw hex of the 2 value bytes, NOT EBCDIC text - a naive
            # printable-text heuristic misfires here since those same
            # bytes happen to also decode to printable EBCDIC chars).
            # F123/F127 unconfirmed; hex is the safer default.
            renderer = lambda ds, tag, val: render_hex(tag, val)
        datasets = decode_dataset_tlv(body, value_renderer=renderer)
        parts = []
        for ds_id, tlvs in datasets:
            for tag, val in tlvs.items():
                parts.append(f"F{fnum}.{ds_id}.{tag})" + val)
        return (SUBFIELD_SEP.join(parts) + SUBFIELD_SEP if parts else ""), pos2 + length

    if t == "field126":
        return decode_field_126(data, pos)

    # fallback
    length, pos2 = read_len_prefix(data, pos)
    val = to_hex(data[pos2:pos2 + length])
    return val, pos2 + length


FIELD126_SUBFIELDS = {8: 20, 9: 20, 10: 6}  # confirmed lengths; others unconfirmed


def decode_field_126(data: bytes, pos: int):
    length, pos2 = read_len_prefix(data, pos)  # 1-byte length = byte count (private-use-style)
    body = data[pos2:pos2 + length]
    bitmap = body[:8]
    rest = body[8:]
    bits = []
    for byte in bitmap:
        for i in range(8):
            bits.append((byte >> (7 - i)) & 1)
    parts = []
    rpos = 0
    for sub in range(1, 65):
        if sub <= len(bits) and bits[sub - 1]:
            sublen = FIELD126_SUBFIELDS.get(sub)
            if sublen is None:
                # Unconfirmed subfield length - consume nothing further,
                # flag it rather than silently misreading later subfields.
                parts.append(f"F126.{sub})<unconfirmed-subfield-length>")
                continue
            subdata = rest[rpos:rpos + sublen]
            rpos += sublen
            if sub == 10:
                # Confirmed: subfield 10 is EBCDIC text where a literal
                # 'F' character is the switch's masking placeholder
                # (mirrors the 0xF-nibble = masked-digit convention used
                # elsewhere), so 'F'/'f' render as '*'.
                subval = decode_ebcdic(subdata).replace("F", "*").replace("f", "*")
            else:
                subval = to_hex(subdata)  # confirmed hex passthrough for 8/9
            parts.append(f"F126.{sub})" + subval)
    return (SUBFIELD_SEP.join(parts) + SUBFIELD_SEP if parts else ""), pos2 + length


F44_SUBFIELD_LENS = [1, 1, 1, 1, 1, 2, 1, 1, 1, 1, 2, 1, 1, 4, 4]  # subfields 1..15


# ---------------------------------------------------------------------------
# Description dictionaries (for the .txt / .xlsx report and colored console)
# ---------------------------------------------------------------------------
# Confidence varies by entry: header + top-level field descriptions and the
# well-known EMV tags are taken directly from the spec / EMV Book 3 and are
# reliable. Field 34/62/63/126 subfield names beyond the ones already
# confirmed against live samples are best-effort labels, not verified -
# treat them the way we've been treating other unconfirmed items in this
# parser (a starting point to correct once more samples come in).

HEADER_DESCRIPTIONS = {
    "H1": "Header Length",
    "H2": "Header Flag / Format",
    "H3": "Text Format Indicator",
    "H4": "Total Message Length",
    "H5": "Destination Station ID",
    "H6": "Source Station ID",
    "H7": "Round-Trip Control Information",
    "H8": "V.I.P. Flags",
    "H9": "Message Status Flags",
    "H10": "Batch Number",
    "H11": "Reserved",
    "H12": "User Information",
    "H13": "Reject Bitmap",
    "H14": "Reject Data Group",
    "MTI": "Message Type Indicator",
}

FIELD_DESCRIPTIONS = {
    2: "Primary Account Number (PAN)",
    3: "Processing Code",
    4: "Amount, Transaction",
    5: "Amount, Reconciliation",
    6: "Amount, Cardholder Billing",
    7: "Transmission Date and Time",
    8: "Amount, Cardholder Billing Fee",
    9: "Conversion Rate, Reconciliation",
    10: "Conversion Rate, Cardholder Billing",
    11: "System Trace Audit Number (STAN)",
    12: "Time, Local Transaction",
    13: "Date, Local Transaction",
    14: "Date, Expiration",
    15: "Date, Settlement",
    16: "Date, Conversion",
    17: "Date, Capture",
    18: "Merchant Type / Category Code (MCC)",
    19: "Acquiring Institution Country Code",
    20: "PAN Extended, Country Code",
    22: "Point of Service Entry Mode",
    23: "Card Sequence Number",
    24: "Network International Identifier (NII)",
    25: "Point of Service Condition Code",
    26: "Point of Service PIN Capture Code",
    28: "Amount, Transaction Fee",
    32: "Acquiring Institution ID Code",
    33: "Forwarding Institution ID Code",
    34: "Electronic Commerce / Acceptance Environment Data",
    35: "Track 2 Data",
    36: "Track 3 Data",
    37: "Retrieval Reference Number (RRN)",
    38: "Authorization ID Response",
    39: "Response Code",
    41: "Card Acceptor Terminal ID",
    42: "Card Acceptor ID Code",
    43: "Card Acceptor Name/Location",
    44: "Additional Response Data",
    45: "Track 1 Data",
    46: "Amounts, Fees",
    47: "Additional Data - National",
    48: "Additional Data - Private",
    49: "Currency Code, Transaction",
    50: "Currency Code, Reconciliation",
    51: "Currency Code, Cardholder Billing",
    52: "PIN Data",
    53: "Security-Related Control Information",
    54: "Additional Amounts",
    55: "ICC-Related / EMV Data",
    56: "PAR / Reserved ISO",
    57: "Reserved National",
    58: "Reserved National",
    59: "National POS Geographic Data",
    60: "Additional POS Information (private-use)",
    61: "Other Amounts (private-use)",
    62: "CPS / Custom Payment Service Fields",
    63: "Visa Private-Use Fields",
    66: "Settlement Code",
    67: "Extended Payment Code",
    68: "Receiving Institution Country Code",
    69: "Settlement Institution Country Code",
    70: "Network Management Information Code",
    90: "Original Data Elements",
    100: "Receiving Institution ID Code",
    102: "Account Identification 1",
    103: "Account Identification 2",
    104: "Transaction Description Data",
    111: "Amount / Currency Reference Data",
    123: "Additional Data 2",
    124: "Additional Data 3",
    125: "Additional Data 4",
    126: "VSDC / Issuer Script Data",
    127: "File Maintenance / Network Data",
    134: "Visa Discretionary Data",
    142: "Issuer Script",
}

# EMV tags seen in Field 55 (and reused for context in F34's dataset '56').
EMV_TAG_NAMES = {
    "82": "Application Interchange Profile (AIP)",
    "84": "Dedicated File Name (AID)",
    "95": "Terminal Verification Results (TVR)",
    "9a": "Transaction Date",
    "9c": "Transaction Type",
    "5f2a": "Transaction Currency Code",
    "5f34": "Application PAN Sequence Number",
    "9f02": "Amount, Authorized (Numeric)",
    "9f03": "Amount, Other (Numeric)",
    "9f10": "Issuer Application Data (IAD)",
    "9f1a": "Terminal Country Code",
    "9f26": "Application Cryptogram (AC)",
    "9f27": "Cryptogram Information Data (CID)",
    "9f33": "Terminal Capabilities",
    "9f34": "Cardholder Verification Method (CVM) Results",
    "9f35": "Terminal Type",
    "9f36": "Application Transaction Counter (ATC)",
    "9f37": "Unpredictable Number",
    "9f41": "Transaction Sequence Counter",
    "9f6e": "Form Factor Indicator / Third-Party Data",
    "91": "Issuer Authentication Data",
    "71": "Issuer Script Template 1",
    "72": "Issuer Script Template 2",
    "86": "Issuer Script Command",
}

# Field 34 dataset '01' subfield tags - confirmed meanings from live samples
# (terminal/app version string and IP address); dataset '56' tags overlap
# with EMV tag numbers but carry different (proprietary/risk) values here -
# names below for dataset 56 are best-effort, not spec-confirmed.
FIELD34_DATASET01_NAMES = {
    "86": "Application/Software Version",
    "89": "Terminal IP Address",
    "c0": "Proprietary Indicator",
}
FIELD34_DATASET56_NAMES = {
    "9f28": "Proprietary Risk Parameter (unconfirmed)",
    "9f29": "Proprietary Risk Parameter (unconfirmed)",
    "9f20": "Proprietary Risk Parameter (unconfirmed)",
}

# Field 44 subfields - best-effort labels; .1-.13 lengths are confirmed
# against live samples, exact names beyond that are indicative only.
FIELD44_DESCRIPTIONS = {
    1: "Response Source / Reason Code",
    2: "Address Verification Service (AVS) Result",
    3: "Additional Token Response Info",
    4: "Extended STIP Reason Code",
    5: "CVV / iCVV / CVV2 Result",
    6: "PACM Diversion-Level Code",
    7: "PACM Diversion Reason Code",
    8: "Card Authentication Results Code",
    9: "Reserved",
    10: "CVV2 Result Code",
    11: "Original Response Code",
    12: "Reserved",
    13: "CAVV Result Code",
    14: "Response Reason Code",
    15: "Reserved",
}

# Field 62 (CPS) subfields - lengths confirmed via spec table; names below
# for the ones not yet seen in live traffic are best-effort.
FIELD62_DESCRIPTIONS = {
    1: "CPS Program Indicator",
    2: "Transaction Identifier",
    3: "Validation Code",
    4: "Market-Specific Data Identifier",
    5: "Prepaid Card Indicator",
    6: "Service Development Field",
    7: "Additional Merchant Data",
    8: "Merchant Data (unconfirmed)",
    9: "Reserved",
    10: "Merchant/Transaction Data (unconfirmed)",
    11: "Reserved",
    12: "Reserved",
    13: "Reserved",
    14: "Merchant Data (unconfirmed)",
    15: "Reserved",
    16: "Reserved",
    17: "Merchant Data (unconfirmed)",
    18: "Reserved",
    19: "Reserved",
    20: "Reserved",
    21: "Reserved",
    22: "Merchant Data (unconfirmed)",
    23: "Reserved",
    24: "Merchant Data (unconfirmed)",
    25: "Reserved",
}

# Field 63 (Visa private-use) subfields - best-effort labels.
FIELD63_DESCRIPTIONS = {
    1: "Settlement/Processing Indicator",
    2: "Reserved",
    3: "Reserved",
    4: "Reserved",
    6: "Reserved",
    7: "Reserved",
    8: "Reserved",
    9: "Reserved",
    10: "Reserved",
    11: "Reserved",
    12: "Reserved",
    13: "Reserved",
    14: "Reserved",
    15: "Reserved",
    18: "Reserved",
    19: "Reserved",
    21: "Reserved",
}

FIELD126_DESCRIPTIONS = {
    8: "Issuer Script Data 1 (encrypted, unconfirmed exact semantics)",
    9: "Issuer Script Data 2 (unconfirmed exact semantics)",
    10: "Issuer Script Results / Status",
}


def describe_label(label: str) -> str:
    """
    Resolve a human-readable description for a compact-format label such
    as 'H1', 'F2', 'F34.01.86', 'F55.9f27', 'F62.2', 'F126.10'.
    """
    if label.startswith("H"):
        return HEADER_DESCRIPTIONS.get(label, "Header field")

    parts = label[1:].split(".")
    try:
        fnum = int(parts[0])
    except ValueError:
        return "Unknown field"

    base_desc = FIELD_DESCRIPTIONS.get(fnum, "Unrecognized / proprietary field")

    if len(parts) == 1:
        return base_desc

    if fnum == 55 and len(parts) == 2:
        tag = parts[1].lower()
        return f"{base_desc} - {EMV_TAG_NAMES.get(tag, 'Proprietary/unrecognized EMV tag ' + tag.upper())}"

    if fnum == 34 and len(parts) == 3:
        dataset, tag = parts[1].lower(), parts[2].lower()
        if dataset == "01":
            return f"{base_desc} - {FIELD34_DATASET01_NAMES.get(tag, 'Unrecognized subfield ' + tag.upper())}"
        if dataset == "56":
            return f"{base_desc} - {FIELD34_DATASET56_NAMES.get(tag, 'Unrecognized subfield ' + tag.upper())}"
        return f"{base_desc} - Dataset {dataset.upper()}, tag {tag.upper()} (unconfirmed)"

    if fnum in (104, 111, 123, 127) and len(parts) == 3:
        dataset, tag = parts[1], parts[2].lower()
        return f"{base_desc} - Dataset {dataset}, {EMV_TAG_NAMES.get(tag, 'tag ' + tag.upper())}"

    if fnum == 44 and len(parts) == 2:
        sub = int(parts[1])
        return f"{base_desc} - Subfield {sub}: {FIELD44_DESCRIPTIONS.get(sub, 'Reserved')}"

    if fnum == 62 and len(parts) == 2:
        sub = int(parts[1])
        return f"{base_desc} - Subfield {sub}: {FIELD62_DESCRIPTIONS.get(sub, 'Reserved')}"

    if fnum == 63 and len(parts) == 2:
        sub = int(parts[1])
        return f"{base_desc} - Subfield {sub}: {FIELD63_DESCRIPTIONS.get(sub, 'Reserved')}"

    if fnum == 126 and len(parts) == 2:
        sub = int(parts[1])
        return f"{base_desc} - Subfield {sub}: {FIELD126_DESCRIPTIONS.get(sub, 'Reserved')}"

    return base_desc


def decode_field_44(data: bytes, pos: int):
    length, pos2 = read_len_prefix(data, pos)
    body = data[pos2:pos2 + length]
    parts = []
    rpos = 0
    for i, sublen in enumerate(F44_SUBFIELD_LENS, start=1):
        if rpos + sublen <= len(body):
            subval = decode_ebcdic_display(body[rpos:rpos + sublen])
            rpos += sublen
        else:
            subval = ""  # ran out of bytes - subfield not present in this instance
        parts.append(f"F44.{i})" + subval)
    return (SUBFIELD_SEP.join(parts) + SUBFIELD_SEP if parts else ""), pos2 + length


def decode_field_62(data: bytes, pos: int):
    length, pos2 = read_len_prefix(data, pos)
    body = data[pos2:pos2 + length]
    bitmap = body[:8]
    rest = body[8:]
    bits = []
    for byte in bitmap:
        for i in range(8):
            bits.append((byte >> (7 - i)) & 1)
    parts = []
    rpos = 0
    for sub in range(1, 65):
        if sub <= len(bits) and bits[sub - 1]:
            sublen = FIELD62_SUBFIELDS.get(sub)
            if sublen is None:
                continue
            subdata = rest[rpos:rpos + sublen]
            rpos += sublen
            if FIELD62_TYPE.get(sub) == "bcd":
                subval = decode_bcd_full(subdata)
            else:
                subval = decode_ebcdic(subdata)
            parts.append(f"F62.{sub})" + subval)
    return (SUBFIELD_SEP.join(parts) + SUBFIELD_SEP if parts else ""), pos2 + length


def decode_field_63(data: bytes, pos: int):
    length, pos2 = read_len_prefix(data, pos)
    body = data[pos2:pos2 + length]
    bitmap = body[:3]
    rest = body[3:]
    bits = []
    for byte in bitmap:
        for i in range(8):
            bits.append((byte >> (7 - i)) & 1)
    parts = []
    rpos = 0
    for sub in range(1, 25):
        if sub <= len(bits) and bits[sub - 1]:
            sublen = FIELD63_SUBFIELDS.get(sub)
            if sublen is None:
                continue
            subdata = rest[rpos:rpos + sublen]
            rpos += sublen
            if FIELD63_TYPE.get(sub) == "bcd":
                subval = decode_bcd_full(subdata)
            else:
                subval = decode_ebcdic(subdata)
            parts.append(f"F63.{sub})" + subval)
    return (SUBFIELD_SEP.join(parts) + SUBFIELD_SEP if parts else ""), pos2 + length


# ---------------------------------------------------------------------------
# Top level message parser
# ---------------------------------------------------------------------------

def _split_compact_field(fnum: int, rendered: str, raw_hex: str):
    """
    Given the compact rendering of one top-level field (which may contain
    several 'Label)value' groups terminated by SUBFIELD_SEP for composite
    fields), return a list of (label, value) row tuples. Simple fields
    return a single row; composite fields (34/44/55/62/63/104/111/126)
    return one row per subfield/tag.
    """
    rows = []
    if fnum in (34, 44, 55, 62, 63, 104, 111, 123, 126, 127):
        groups = [g for g in rendered.split(SUBFIELD_SEP) if g]
        for g in groups:
            if ")" in g:
                label, value = g.split(")", 1)
            else:
                label, value = f"F{fnum}", g
            rows.append((label, value))
        if not rows:
            rows.append((f"F{fnum}", ""))
    else:
        rows.append((f"F{fnum}", rendered))
    return rows


def parse_message_full(raw_hex: str, debug: bool = False):
    """
    Full parse returning both the compact string (identical to
    parse_message's output) and a list of detailed rows for reporting:
    each row is (label, description, raw_hex, value).
    """
    data = clean_hex(raw_hex)
    header, pos = parse_header(data)
    mti = decode_bcd_full(data[pos:pos + 2])
    mti_start = pos
    pos += 2
    present_fields, pos = parse_all_bitmaps(data, pos)
    if debug:
        print(f"[debug] header_len={header['_header_len']} mti_end_pos={pos}", file=sys.stderr)

    rows = []

    # Header rows (byte ranges recomputed from known fixed sizes)
    header_layout = [("H1", 1), ("H2", 1), ("H3", 1), ("H4", 2), ("H5", 3),
                      ("H6", 3), ("H7", 1), ("H8", 2), ("H9", 3), ("H10", 1),
                      ("H11", 3), ("H12", 1)]
    if header["_is_reject"]:
        header_layout += [("H13", 2), ("H14", 2)]
    hpos = 0
    for key, size in header_layout:
        rows.append((key, describe_label(key), data[hpos:hpos + size].hex().upper(), str(header[key])))
        hpos += size
    rows.append(("MTI", describe_label("MTI"), data[mti_start:mti_start + 2].hex().upper(), mti))

    out_parts = []
    for key in ["H1", "H2", "H3", "H4", "H5", "H6", "H7", "H8", "H9", "H10", "H11", "H12"]:
        out_parts.append(f"{key})" + str(header[key]))
    if header["_is_reject"]:
        out_parts.append("H13)" + str(header["H13"]))
        out_parts.append("H14)" + str(header["H14"]))

    for fnum in sorted(present_fields):
        start = pos
        try:
            rendered, pos = decode_field(fnum, data, pos)
        except Exception as e:
            rendered = f"<parse-error: {e}>"
        if debug:
            print(f"[debug] F{fnum}: bytes[{start}:{pos}] = {data[start:pos].hex()} -> {rendered!r}", file=sys.stderr)

        raw_hex_field = data[start:pos].hex().upper()
        for label, value in _split_compact_field(fnum, rendered, raw_hex_field):
            rows.append((label, describe_label(label), raw_hex_field, value))

        if fnum in (44, 62, 63, 126) or fnum in (34, 55, 104, 111, 123, 127):
            out_parts.append(rendered)  # already contains F{n}.{sub}) formatting
        else:
            out_parts.append(f"F{fnum})" + rendered)

    compact = f"{mti[1:] if mti.startswith('0') else mti}: " + FIELD_SEP.join(out_parts) + FIELD_SEP
    return compact, rows


def parse_message(raw_hex: str, debug: bool = False) -> str:
    return parse_message_full(raw_hex, debug=debug)[0]


# ---------------------------------------------------------------------------
# Report export: colored console, .txt, .xlsx (mirrors the TIC-style output
# convention: one folder per message named <RRN_or_STAN>_<MTI>, containing
# a .txt and a .xlsx report, plus a colored summary printed to console)
# ---------------------------------------------------------------------------

class _Ansi:
    RESET = "\033[0m"
    BOLD = "\033[1m"
    CYAN = "\033[96m"
    YELLOW = "\033[93m"
    GREEN = "\033[92m"
    MAGENTA = "\033[95m"
    RED = "\033[91m"
    DIM = "\033[2m"


def _supports_color() -> bool:
    if sys.platform == "win32":
        try:
            import colorama  # noqa: F401
            colorama.just_fix_windows_console()
            return True
        except Exception:
            return False
    return sys.stdout.isatty()


def print_colored_report(compact: str, rows, mti: str):
    use_color = _supports_color()

    def c(text, color):
        return f"{color}{text}{_Ansi.RESET}" if use_color else text

    print()
    print(c(f"===== VisaNet ISO8583 Message (MTI {mti}) =====", _Ansi.BOLD + _Ansi.CYAN))
    for label, desc, raw_hex, value in rows:
        label_c = c(f"{label:<14}", _Ansi.YELLOW)
        desc_c = c(f"{desc:<55}", _Ansi.DIM)
        value_c = c(value, _Ansi.GREEN)
        print(f"  {label_c} {desc_c} {value_c}")
    print()


def get_report_name(rows, mti: str) -> str:
    """Folder/file base name: <RRN_or_STAN>_<MTI>, mirroring the TIC
    convention of D:\\TIC_Output\\<RRN>_<MTI>\\."""
    row_map = {label: value for label, _desc, _raw, value in rows}
    rrn = row_map.get("F37", "").strip()
    stan = row_map.get("F11", "").strip()
    key = rrn if rrn else (stan if stan else "UNKNOWN")
    return f"{key}_{mti}"

def write_txt_report(path: str, raw_hex: str, compact: str, rows):
    from datetime import datetime

    with open(path, "w", encoding="utf-8") as f:
        f.write(f"Generated: {datetime.now().isoformat(timespec='seconds')}\n\n")

        # RAW MESSAGE
        f.write("=" * 120 + "\n")
        f.write("RAW MESSAGE (EXACT INPUT)\n")
        f.write("=" * 120 + "\n")
        f.write(raw_hex)
        f.write("\n\n")

        # COMPACT
        f.write("=" * 120 + "\n")
        f.write("COMPACT PARSED MESSAGE\n")
        f.write("=" * 120 + "\n")
        f.write(compact)
        f.write("\n\n")

        # FIELD BREAKDOWN
        f.write("=" * 120 + "\n")
        f.write("FIELD-BY-FIELD BREAKDOWN\n")
        f.write("=" * 120 + "\n")

        field_width = max(
            len("Field"),
            max(len(str(label)) for label, _, _, _ in rows)
        ) + 2

        desc_width = max(
            len("Description"),
            max(len(str(desc)) for _, desc, _, _ in rows)
        ) + 2

        value_width = max(
            len("Value"),
            max(len(str(value)) for _, _, _, value in rows)
        ) + 2

        header = (
            f"{'Field':<{field_width}}"
            f"{'Description':<{desc_width}}"
            f"{'Value'}"
        )

        f.write(header + "\n")
        f.write("-" * len(header) + "\n")

        for label, desc, raw, value in rows:
            value = "" if value is None else str(value)

            f.write(
                f"{label:<{field_width}}"
                f"{desc:<{desc_width}}"
                f"{value}\n"
            )

def write_xlsx_report(path: str, raw_hex: str, compact: str, rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl is not installed - skipping .xlsx export "
              "(install with: pip install openpyxl --break-system-packages)")
        return

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = "Raw Message (hex)"
    ws_summary["A1"].font = Font(bold=True)
    ws_summary["A2"] = raw_hex
    ws_summary["A2"].alignment = Alignment(wrap_text=True)
    ws_summary["A4"] = "Compact Parsed Message"
    ws_summary["A4"].font = Font(bold=True)
    ws_summary["A5"] = compact
    ws_summary["A5"].alignment = Alignment(wrap_text=True)
    ws_summary.column_dimensions["A"].width = 120

    ws = wb.create_sheet("Fields")
    headers = ["Field", "Description", "Raw Hex", "Value"]
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for r, (label, desc, raw, value) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=desc)
        ws.cell(row=r, column=3, value=raw)
        ws.cell(row=r, column=4, value=value)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 40
    ws.auto_filter.ref = f"A1:D{len(rows) + 1}"

    wb.save(path)


def export_report(raw_hex_input: str, out_base_dir: str = r"D:\VISA_Output", debug: bool = False):
    """
    Parse the message and write a per-message report folder:
        <out_base_dir>/<RRN_or_STAN>_<MTI>/<name>.txt
        <out_base_dir>/<RRN_or_STAN>_<MTI>/<name>.xlsx
    Also prints a colored field-by-field summary to the console.
    Returns (compact_string, folder_path).
    """
    import os

    cleaned = clean_hex(raw_hex_input)
    compact, rows = parse_message_full(raw_hex_input, debug=debug)
    mti = compact.split(":", 1)[0].strip()
    name = get_report_name(rows, mti)

    folder = os.path.join(out_base_dir, name)
    os.makedirs(folder, exist_ok=True)

    txt_path = os.path.join(folder, f"{name}.txt")
    xlsx_path = os.path.join(folder, f"{name}.xlsx")

    write_txt_report(txt_path, raw_hex_input.strip(), compact, rows)
    write_xlsx_report(
        xlsx_path,
        raw_hex_input.strip(),
        compact,
        rows
    )

    print_colored_report(compact, rows, mti)
    print(f"Saved: {txt_path}")
    print(f"Saved: {xlsx_path}")

    return compact, folder


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    debug = "--debug" in sys.argv
    no_export = "--no-export" in sys.argv
    out_dir = r"D:\VISA_Output"
    for a in sys.argv[1:]:
        if a.startswith("--outdir="):
            out_dir = a.split("=", 1)[1]

    if args:
        # File path given on the command line
        with open(args[0], "r") as f:
            raw = f.read()
    else:
        # Interactive/paste mode. Works whether or not stdin is a real
        # terminal. Paste the hex (single or multiple lines), then press
        # Enter on an empty line to finish.
        print("Paste the raw ISO8583 hex message below.", flush=True)
        print("Press Enter on an empty line when you're done:\n", flush=True)
        lines = []
        try:
            while True:
                line = input()
                if line.strip() == "":
                    break
                lines.append(line)
        except EOFError:
            pass  # also accept Ctrl+D / Ctrl+Z as an alternative way to finish
        raw = "\n".join(lines)

    if not raw.strip():
        print("\nNo input received - nothing to parse.")
        sys.exit(1)

    try:
        clean_hex(raw)  # validate early with a friendly error
    except Exception as e:
        print(f"\nCould not parse input as hex: {e}")
        sys.exit(1)

    try:
        if no_export:
            print("\n=== Parsed message ===")
            print(parse_message(raw, debug=debug))
        else:
            export_report(raw, out_base_dir=out_dir, debug=debug)
    except Exception as e:
        print(f"<parser error: {e}>")